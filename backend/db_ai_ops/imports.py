import csv
import io
import json
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook


def _strip(v):
    if v is None:
        return ''
    return str(v).strip()


def parse_bool(v, default=False):
    s = _strip(v).lower()
    if s == '':
        return default
    if s in {'1', 'true', 'yes', 'y', 'on', '启用', '是'}:
        return True
    if s in {'0', 'false', 'no', 'n', 'off', '禁用', '否'}:
        return False
    return default


def parse_tags(v):
    s = _strip(v)
    if not s:
        return []
    parts = []
    for chunk in s.replace(';', ',').split(','):
        t = chunk.strip()
        if t:
            parts.append(t)
    return parts


def parse_json(v, default=None):
    if default is None:
        default = {}
    s = _strip(v)
    if not s:
        return default
    return json.loads(s)


def normalize_headers(headers: Iterable[str], mapping: Dict[str, str]) -> List[str]:
    out = []
    for h in headers:
        raw = _strip(h)
        key = raw.lower()
        out.append(mapping.get(raw) or mapping.get(key) or key)
    return out


def read_csv_or_txt(file_bytes: bytes) -> Tuple[List[str], List[List[Any]]]:
    text = file_bytes.decode('utf-8-sig', errors='replace')
    buf = io.StringIO(text)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[',', '\t', ';', '|'])
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ','

    reader = csv.reader(buf, dialect)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [_strip(x) for x in rows[0]]
    data = rows[1:]
    return headers, data


def read_xlsx(file_bytes: bytes) -> Tuple[List[str], List[List[Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_strip(x) for x in rows[0]]
    data = [list(r) for r in rows[1:] if any((_strip(c) for c in r))]
    return headers, data


def build_xlsx_template(headers: List[str], example_row: List[Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    ws.append(example_row)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_csv_template(headers: List[str], example_row: List[Any]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerow(example_row)
    text = buf.getvalue()
    return ('\ufeff' + text).encode('utf-8')


def build_tsv_template(headers: List[str], example_row: List[Any]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter='\t', lineterminator='\n')
    writer.writerow(headers)
    writer.writerow(example_row)
    text = buf.getvalue()
    return ('\ufeff' + text).encode('utf-8')


def rows_to_dicts(headers: List[str], rows: List[List[Any]]) -> List[Dict[str, Any]]:
    out = []
    for row in rows:
        item = {}
        for idx, h in enumerate(headers):
            item[h] = row[idx] if idx < len(row) else None
        out.append(item)
    return out
